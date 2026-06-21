"""Vanguard *Full History* XLSX parser (v1.4 — full transaction history workaround).

Vanguard's web "Activity → Download" only exposes the last 18 months of
transactions. The user-facing workaround is to run an "Activity report"
under *Reports* with a custom date range and export it to Excel; that
file is a single-sheet workbook with the layout::

    Settlement date | Trade date | Symbol | Name | Transaction type |
    Account type    | Quantity   | Price  | Commission & fees** | Amount

Differences from the brokerage CSV (:mod:`.parser`):

* The first two rows are a "Custom report created on …" preamble — the
  real header lives a few rows down. We locate it by looking for the
  ``"Settlement date"`` cell.
* ``Amount`` is formatted as a string with a leading ``$``, thousand
  separators and *signed* values (buys negative, sells/dividends
  positive). We rely on those signs and do **not** flip ``Quantity``
  for sells — the XLSX already exports sells with a negative
  ``Quantity``.
* ``Commission & fees`` is the literal string ``"Free"`` for zero-fee
  trades; everything else is a ``$``-prefixed decimal.
* There is no separate ``Principal Amount`` column — gross is derived
  from ``Quantity × Price`` when both are present.
* New transaction types appear in this export: ``Stock split`` and
  ``Funds Received (adjustment)`` (a post-settlement correction to an
  earlier deposit). Both are handled in
  :mod:`.action_map`.

Sweep rows are dropped just like in :mod:`.parser`; the count is
returned alongside the parsed rows.

Numbers/dates are assumed **US-locale** (see
:mod:`investment_dashboard.adapters.locale_parsing`). Per-row problems
(unknown action, un-parseable / EU-locale cell) are *collected* into the
returned :class:`ParseReport` rather than aborting the whole import (audit
D3); a structural header/layout mismatch still aborts loudly (audit D2).
"""

from __future__ import annotations

import hashlib
import io
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook

from investment_dashboard.adapters.importer_types import (
    ParsedTransactionRow,
    ParseReport,
    RowIssue,
    UnknownActionError,
)
from investment_dashboard.adapters.locale_parsing import (
    LocaleError,
    parse_us_date,
    parse_us_decimal,
)
from investment_dashboard.adapters.row_validation import validate_row
from investment_dashboard.adapters.vanguard.action_map import map_transaction_type

# Canonical header tokens (lower-cased, ``**`` footnote markers stripped).
_REQUIRED_HEADERS = ("settlement date", "trade date", "transaction type", "amount")


@dataclass(frozen=True)
class VanguardXlsxParseResult(ParseReport):
    """Backwards-compatible name for :class:`ParseReport`.

    Retained so existing call sites/tests that import
    ``VanguardXlsxParseResult`` keep working; carries the new
    ``unknown_actions`` / ``errors`` / ``warnings`` fields.
    """


def _normalise_header(h: object) -> str:
    s = "" if h is None else str(h)
    return s.strip().lower().rstrip("*").strip()


def _to_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_date(v: object) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return parse_us_date(_to_str(v))


def _parse_decimal(v: object) -> Decimal | None:
    """Parse a Vanguard XLSX numeric cell.

    Tolerates raw numbers, ``"$1,234.5600"``, ``"-$10,000.0000"``,
    ``"Free"``, blanks and dashes. EU-locale strings raise
    :class:`LocaleError` via :func:`parse_us_decimal`.
    """
    if v is None:
        return None
    if isinstance(v, int | float | Decimal):
        return Decimal(str(v))
    return parse_us_decimal(str(v))


def _hash_external_id(*parts: object) -> str:
    payload = "|".join(str(p) for p in parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:32]


def _find_header_row(grid: list[list[object]]) -> int:
    for i, row in enumerate(grid):
        lowered = [_normalise_header(c) for c in row]
        if all(any(want == cell for cell in lowered) for want in _REQUIRED_HEADERS):
            return i
    raise ValueError(
        "Vanguard XLSX: could not locate header row containing "
        "'Settlement date', 'Trade date', 'Transaction type' and 'Amount'"
    )


def _row_records(
    grid: list[list[object]], header_row: int
) -> Iterable[tuple[int, dict[str, object]]]:
    headers = [_normalise_header(c) for c in grid[header_row]]
    for offset, raw in enumerate(grid[header_row + 1 :], start=header_row + 2):
        # ``offset`` is the 1-based spreadsheet row number.
        # Skip fully empty rows / trailing disclaimer text that openpyxl
        # returns as a single non-empty cell.
        if not any(_to_str(c) for c in raw):
            continue
        # A populated cell that maps to no header (a blank/short header column)
        # would be silently dropped by the ``if key`` filter below. That means
        # the sheet layout doesn't line up with the detected header, so refuse
        # rather than truncate data. Trailing *empty* cells (openpyxl pads
        # ragged rows to the widest row) are fine.
        for idx, value in enumerate(raw):
            key = headers[idx] if idx < len(headers) else ""
            if not key and _to_str(value):
                raise ValueError(
                    "Vanguard XLSX: a data cell falls under a column with no "
                    f"header (column {idx + 1}); the file layout doesn't match "
                    "the detected header. Refusing to import rather than "
                    "silently drop data."
                )
        record: dict[str, object] = {}
        for key, value in zip(headers, raw, strict=False):
            if key:
                record[key] = value
        if not _to_str(record.get("transaction type")):
            # Footnote rows ("** Vanguard Brokerage Services® does not …")
            # — they live below the data and lack a transaction type.
            continue
        yield offset, record


def _parse_record(record: dict[str, object]) -> ParsedTransactionRow | None:
    """Parse one record. Returns ``None`` for a sweep; raises on a bad cell."""
    txn_type = _to_str(record.get("transaction type"))
    kind = map_transaction_type(txn_type)
    if kind is None:
        return None  # sweep — caller counts it

    trade_date = _parse_date(record.get("trade date") or record.get("settlement date"))
    settlement_raw = record.get("settlement date")
    settlement = _parse_date(settlement_raw) if _to_str(settlement_raw) else None

    symbol = _to_str(record.get("symbol")) or None
    quantity = _parse_decimal(record.get("quantity"))
    price = _parse_decimal(record.get("price"))
    commission = _parse_decimal(record.get("commission & fees")) or Decimal(0)
    net = _parse_decimal(record.get("amount"))
    name = _to_str(record.get("name")) or None

    # Vanguard's XLSX export *already* signs sell quantities negative,
    # so unlike the CSV path we do not flip them here.

    gross = quantity * price if (quantity is not None and price is not None) else None

    external_id = _hash_external_id(
        trade_date.isoformat(),
        txn_type,
        symbol or "",
        quantity if quantity is not None else "",
        price if price is not None else "",
        net if net is not None else "",
    )

    return ParsedTransactionRow(
        date=trade_date,
        settlement_date=settlement,
        kind=kind,
        symbol=symbol,
        quantity=quantity,
        price_native=price,
        gross_native=gross,
        fees_native=commission if commission != 0 else None,
        net_native=net,
        description=name,
        external_id=external_id,
        source="import_vanguard_xlsx",
        name=name or None,
    )


def parse_vanguard_xlsx(content: bytes) -> VanguardXlsxParseResult:
    """Parse a Vanguard Full-History Excel workbook into a :class:`ParseReport`.

    Unknown transaction types and bad/EU-locale cells are collected as
    per-row errors (the row is skipped) instead of aborting; a structural
    header/layout mismatch still raises loudly. Light consistency-check
    failures are reported as warnings.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb.active
        grid: list[list[object]] = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not grid:
        return VanguardXlsxParseResult(rows=[])

    header_row = _find_header_row(grid)

    rows: list[ParsedTransactionRow] = []
    sweeps_dropped = 0
    unknown_actions: list[str] = []
    errors: list[RowIssue] = []
    warnings: list[RowIssue] = []

    for line, record in _row_records(grid, header_row):
        txn_type = _to_str(record.get("transaction type"))
        try:
            parsed = _parse_record(record)
        except UnknownActionError as exc:
            unknown_actions.append(txn_type)
            errors.append(RowIssue(message=str(exc), line=line, raw=txn_type))
            continue
        except (LocaleError, ValueError) as exc:
            errors.append(RowIssue(message=str(exc), line=line, raw=txn_type))
            continue
        if parsed is None:
            sweeps_dropped += 1
            continue
        rows.append(parsed)
        for warn in validate_row(parsed):
            warnings.append(RowIssue(message=warn, line=line, raw=txn_type))

    return VanguardXlsxParseResult(
        rows=rows,
        sweeps_dropped=sweeps_dropped,
        unknown_actions=unknown_actions,
        errors=errors,
        warnings=warnings,
    )
